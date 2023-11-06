# Importamos las librerias
import sqlite3
import random
import string
import re
import datetime
from datetime import datetime
import csv
import openpyxl
import os

# Conexión a la base de datos SQLite

# Función para crear las tablas en sqlite al iniciar el programa
def crear_tablas():
    try:
        with sqlite3.connect("taller.db") as conn:
            cursor = conn.cursor()
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS Clientes (
                idcliente INTEGER PRIMARY KEY AUTOINCREMENT,
                nombrecliente TEXT NOT NULL,
                rfccliente TEXT NOT NULL,
                correocliente TEXT NOT NULL
            )
            """)

            cursor.execute("""
            CREATE TABLE IF NOT EXISTS Servicios (
                idservicio INTEGER PRIMARY KEY AUTOINCREMENT,
                nombreservicio TEXT NOT NULL,
                costoservicio REAL CHECK (costoservicio > 0.00) NOT NULL
            )
            """)

            cursor.execute("""
            CREATE TABLE IF NOT EXISTS Notas (
                folionota INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha TEXT,
                idcliente INTEGER NOT NULL,
                idservicio INTEGER NOT NULL,
                status INTEGER DEFAULT 1,
                FOREIGN KEY (idcliente) REFERENCES Clientes(idcliente),
                FOREIGN KEY (idservicio) REFERENCES Servicios(idservicio)
            )
            """)
    except sqlite3.Error as e:
        print(f"Error al crear tablas en la base de datos: {e}")

# Creamos los submenus
def mostrar_menu_notas():
    print("\nMenú Notas:")
    print("1. Registrar una nota")
    print("2. Cancelar una nota")
    print("3. Recuperar una nota")
    print("4. Consultas y reportes de notas")
    print("5. Volver al menú principal")

def mostrar_menu_clientes():
    print("\nMenú Clientes:")
    print("1. Agregar un cliente")
    print("2. Consultas y reportes de clientes")
    print("3. Volver al menú principal")

def mostrar_menu_servicios():
    print("\nMenú Servicios")
    print("1. Agregar un servicio")
    print("2. Consultas y reportes de servicios")
    print("3. Volver al menú anterior")

# Creamos la opción Notas
def Notas():
    while True:
        mostrar_menu_notas()
        opcion = input("Seleccione una opción: ")

        if opcion == "1":
            registrar_nota()
        elif opcion == "2":
            cancelar_nota()
        elif opcion == "3":
            recuperar_nota()
        elif opcion == "4":
            consultasreportes_notas()
        elif opcion == "5":
            print("Volviendo al menú principal.")
            break
        else:
            print("Opción no válida. Por favor, seleccione una opción válida.")
            
def registrar_nota():
    while True:
        while True:
            fecha_ingresada = input("Ingresa una fecha en el formato DD/MM/YYYY: ")
            if validar_fecha(fecha_ingresada):
                break
            else:
                print("Fecha no válida o es posterior a la fecha actual. Intenta de nuevo.")

        with sqlite3.connect("taller.db") as conn:
            cursor = conn.cursor()
            mostrar_registros("Clientes")
            mostrar_registros("Servicios")

            while True:
                try:
                    cliente = input("\nIngrese la clave del cliente: ")
                    cliente = int(cliente)
                    cursor.execute("SELECT idcliente, nombrecliente FROM Clientes WHERE idcliente = ?", (cliente,))
                    cliente_existente = cursor.fetchone()
                    if cliente_existente:
                        print(f"Cliente seleccionado: {cliente_existente[1]}")
                        break
                    else:
                        print("Clave de cliente no válida. Debe ser mayor que 0 y existir en la tabla Clientes. Intente de nuevo.")
                except ValueError:
                    print("Error: Ingrese una clave de cliente válida.")

            while True:
                try:
                    servicio = int(input("Ingrese la clave del servicio: "))
                    cursor.execute("SELECT idservicio, nombreservicio FROM Servicios WHERE idservicio = ?", (servicio,))
                    servicio_existente = cursor.fetchone()
                    if servicio_existente:
                        print(f"Servicio seleccionado: {servicio_existente[1]}")
                        break
                    else:
                        print("Clave de servicio no válida. Debe ser mayor que 0 y existir en la tabla Servicios. Intente de nuevo.")
                except ValueError:
                    print("Error: Ingrese una clave de servicio válida.")

            # Realizar la inserción de la nota en la base de datos
            try:
                cursor.execute("INSERT INTO Notas (fecha, idcliente, idservicio) VALUES (?, ?, ?)", (fecha_ingresada, cliente, servicio))
                conn.commit()
                cursor.execute("SELECT folionota FROM Notas WHERE folionota = (SELECT MAX(folionota) FROM Notas)")
                folio = cursor.fetchone()
                print(f"Nota agregada exitosamente con folio {folio[0]}")
                break
            except Exception as e:
                print(f"Error al agregar la nota: {e}")

def cancelar_nota():
    while True:
        try:
            folio = int(input("Ingrese el folio de la nota que desea cancelar: "))
        except ValueError:
            print("Error: El folio debe ser un número válido.")
            continue

        with sqlite3.connect("taller.db") as conn:
            cursor = conn.cursor()

            # Verificar si la nota con el folio proporcionado existe y no está cancelada
            cursor.execute("SELECT * FROM Notas WHERE folionota = ? AND status = 1", (folio,))
            nota = cursor.fetchone()

            if nota:
                # Mostrar los detalles de la nota antes de la cancelación
                print("Detalle de la nota a cancelar:")
                print(f"Folio de la nota: {nota[0]}")
                print(f"Fecha: {nota[1]}")
                cursor.execute("SELECT nombrecliente FROM Clientes WHERE idcliente = ?", (nota[2],))
                nombre_cliente = cursor.fetchone()[0]
                cursor.execute("SELECT nombreservicio FROM Servicios WHERE idservicio = ?", (nota[3],))
                nombre_servicio = cursor.fetchone()[0]
                print(f"Cliente: {nombre_cliente}")
                print(f"Servicio: {nombre_servicio}")

                confirmacion = input("¿Está seguro de que desea cancelar esta nota? (S/N): ").strip().upper()

                if confirmacion == 'S':
                    # Cambiar el estado de la nota a cancelada
                    cursor.execute("UPDATE Notas SET status = 0 WHERE folionota = ?", (folio,))
                    conn.commit()
                    print("La nota ha sido cancelada exitosamente.")
                else:
                    print("La cancelación de la nota ha sido cancelada a solicitud del usuario.")
            else:
                print("La nota no existe o ya está cancelada en el sistema.")

        break


def recuperar_nota():
    with sqlite3.connect("taller.db") as conn:
        cursor = conn.cursor()

        # Obtener una lista de notas canceladas (sin su detalle)
        cursor.execute("SELECT folionota FROM Notas WHERE status = 0")
        notas_canceladas = cursor.fetchall()

        if not notas_canceladas:
            print("No hay notas canceladas para recuperar.")
            return

        print("Listado de notas canceladas:")
        for folio in notas_canceladas:
            print(f"Folio de la nota: {folio[0]}")

        while True:
            try:
                folio_recuperar = int(input("Ingrese el folio de la nota que desea recuperar o 0 para cancelar: "))
            except ValueError:
                print("Error: El folio debe ser un número válido.")
                continue

            if folio_recuperar == 0:
                print("Operación cancelada. No se recuperó ninguna nota.")
                return
            elif (folio_recuperar,) in notas_canceladas:
                cursor.execute("SELECT * FROM Notas WHERE folionota = ?", (folio_recuperar,))
                nota = cursor.fetchone()

                # Mostrar el detalle de la nota cancelada antes de la recuperación
                print("Detalle de la nota a recuperar:")
                print(f"Folio de la nota: {nota[0]}")
                print(f"Fecha: {nota[1]}")
                cursor.execute("SELECT nombrecliente FROM Clientes WHERE idcliente = ?", (nota[2],))
                nombre_cliente = cursor.fetchone()[0]
                cursor.execute("SELECT nombreservicio FROM Servicios WHERE idservicio = ?", (nota[3],))
                nombre_servicio = cursor.fetchone()[0]
                print(f"Cliente: {nombre_cliente}")
                print(f"Servicio: {nombre_servicio}")

                confirmacion = input("¿Está seguro de que desea recuperar esta nota? (S/N): ").strip().upper()

                if confirmacion == 'S':
                    # Cambiar el estado de la nota a activa (no cancelada)
                    cursor.execute("UPDATE Notas SET status = 1 WHERE folionota = ?", (folio_recuperar,))
                    conn.commit()
                    print("La nota ha sido recuperada exitosamente.")
                else:
                    print("La recuperación de la nota ha sido cancelada a solicitud del usuario.")
                break
            else:
                print("El folio ingresado no corresponde a una nota cancelada en el sistema.")


def consultasreportes_notas():
    while True:
        print("\nConsultas y Reportes de Notas:")
        print("1. Consulta por período")
        print("2. Consulta por folio")
        print("3. Volver atrás")
        opcion = input("Seleccione una opción: ")

        if opcion == "1":  # Consulta por período
            consulta_por_periodo()
        elif opcion == "2":  # Consulta por folio
            consulta_por_folio()
        elif opcion == "3":  # Volver atrás
            print("Volviendo al menú principal.")
            break
        else:
            print("Opción no válida. Por favor, seleccione una opción válida.")

def consulta_por_periodo():
    while True:
        fecha_inicial = input("Ingrese la fecha inicial (DD-MM-YYYY) o presione Enter para asumir 01/01/2000: ")
        fecha_final = input("Ingrese la fecha final (DD-MM-YYYY) o presione Enter para asumir la fecha actual: ")

        if fecha_inicial == "":
            fecha_inicial = "01-01-2000"
        if fecha_final == "":
            fecha_final = datetime.datetime.now().strftime("%d-%m-%Y")

        try:
            fecha_inicial = datetime.datetime.strptime(fecha_inicial, '%d-%m-%Y')
            fecha_final = datetime.datetime.strptime(fecha_final, '%d-%m-%Y')
        except ValueError:
            print("Error: Las fechas no tienen un formato válido (DD-MM-YYYY).")
            continue

        if fecha_final < fecha_inicial:
            print("Error: La fecha final debe ser igual o posterior a la fecha inicial.")
            continue

        with sqlite3.connect("taller.db") as conn:
            cursor = conn.cursor()

            # Consulta de notas por período
            cursor.execute("""
                SELECT n.folionota, n.fecha, c.nombrecliente
                FROM Notas n
                JOIN Clientes c ON n.idcliente = c.idcliente
                WHERE date(n.fecha) BETWEEN ? AND ?
            """, (fecha_inicial, fecha_final))
            notas = cursor.fetchall()

            if notas:
                # Calcular el monto promedio de las notas del período
                cursor.execute("""
                    SELECT AVG(s.costoservicio)
                    FROM Servicios s
                    JOIN Notas n ON s.idservicio = n.idservicio
                    WHERE date(n.fecha) BETWEEN ? AND ?
                """, (fecha_inicial, fecha_final))
                promedio_monto = cursor.fetchone()[0]

                print("Notas dentro del período seleccionado:")
                print(f"{'Folio':<10} {'Fecha':<12} {'Nombre del Cliente':<30}")
                for nota in notas:
                    print(f"{nota[0]:<10} {nota[1]:<12} {nota[2]:<30}")

                print(f"Monto promedio del período: {promedio_monto:.2f}")

                while True:
                    opcion_exportar = input("¿Desea exportar el reporte a CSV, Excel o regresar al menú de reportes? (CSV/Excel/Regresar): ").strip().lower()
                    if opcion_exportar == "csv":
                        exportar_reporte_csv(notas, fecha_inicial, fecha_final, promedio_monto)
                        break
                    elif opcion_exportar == "excel":
                        exportar_reporte_excel(notas, fecha_inicial, fecha_final, promedio_monto)
                        break
                    elif opcion_exportar == "regresar":
                        break
                    else:
                        print("Opción no válida. Por favor, seleccione una opción válida.")
            else:
                print("No hay notas emitidas para el período especificado.")
            break
def consulta_por_folio():
    while True:
        with sqlite3.connect("taller.db") as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT n.folionota, n.fecha, c.nombrecliente, s.nombreservicio
                FROM Notas n
                JOIN Clientes c ON n.idcliente = c.idcliente
                JOIN Servicios s ON n.idservicio = s.idservicio
                WHERE n.status = 1
                ORDER BY n.folionota
            """)
            notas = cursor.fetchall()

            if not notas:
                print("No hay notas activas en el sistema para consultar.")
                return

            print("Listado de notas activas:")
            print("Folio  |  Fecha       |  Nombre del Cliente  |  Servicio")
            for nota in notas:
                print(f"{nota[0]:<7} | {nota[1]} | {nota[2]} | {nota[3]}")

            try:
                folio_consultar = int(input("Ingrese el folio de la nota que desea consultar o 0 para salir: "))
            except ValueError:
                print("Error: El folio debe ser un número válido.")
                continue

            if folio_consultar == 0:
                print("Saliendo del menú de consulta por folio.")
                break

            if (folio_consultar,) in [(n[0],) for n in notas]:
                cursor.execute("""
                    SELECT n.folionota, n.fecha, c.nombrecliente, c.rfccliente, c.correocliente, s.nombreservicio, s.costoservicio
                    FROM Notas n
                    JOIN Clientes c ON n.idcliente = c.idcliente
                    JOIN Servicios s ON n.idservicio = s.idservicio
                    WHERE n.folionota = ? AND n.status = 1
                """, (folio_consultar,))
                nota = cursor.fetchone()

                print("Detalle de la nota:")
                print(f"Folio de la nota: {nota[0]}")
                print(f"Fecha: {nota[1]}")
                print(f"Nombre del Cliente: {nota[2]}")
                print(f"RFC del Cliente: {nota[3]}")
                print(f"Correo del Cliente: {nota[4]}")
                print(f"Servicio: {nota[5]}")
                print(f"Costo del servicio: {nota[6]:.2f}")
            else:
                print("El folio ingresado no corresponde a una nota activa en el sistema o no existe.")




def exportar_reporte_csv(notas, fecha_inicial, fecha_final, promedio_monto):
    if fecha_inicial is not None and fecha_final is not None:
        filename = f"ReportePorPeriodo_{fecha_inicial.strftime('%m_%d_%Y')}_{fecha_final.strftime('%m_%d_%Y')}.csv"
    else:
        filename = "ReporteNotas.csv"

    with open(filename, mode='w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(["Folio", "Fecha", "Nombre del Cliente"])
        for nota in notas:
            writer.writerow(nota)
        if fecha_inicial is not None and fecha_final is not None:
            writer.writerow(["Monto promedio del período:", "", promedio_monto])

    print(f"El reporte ha sido exportado como {filename}.")

def exportar_reporte_excel(notas, fecha_inicial, fecha_final, promedio_monto):
    if fecha_inicial is not None and fecha_final is not None:
        filename = f"ReportePorPeriodo_{fecha_inicial.strftime('%m_%d_%Y')}_{fecha_final.strftime('%m_%d_%Y')}.xlsx"
    else:
        filename = "ReporteNotas.xlsx"

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Reporte de Notas"

    bold = openpyxl.styles.Font(bold=True)
    bold_style = openpyxl.styles.NamedStyle(name="bold")
    bold_style.font = bold

    worksheet['A1'] = "Folio"
    worksheet['B1'] = "Fecha"
    worksheet['C1'] = "Nombre del Cliente"

    worksheet['A1'].style = bold_style
    worksheet['B1'].style = bold_style
    worksheet['C1'].style = bold_style

    row = 2

    for nota in notas:
        worksheet.cell(row=row, column=1, value=nota[0])
        worksheet.cell(row=row, column=2, value=nota[1])
        worksheet.cell(row=row, column=3, value=nota[2])
        row += 1

    if fecha_inicial is not None and fecha_final is not None:
        worksheet.cell(row=row, column=1, value="Monto promedio del período:")
        worksheet.cell(row=row, column=3, value=promedio_monto)
        worksheet.cell(row=row, column=1).style = bold_style
        worksheet.cell(row=row, column=3).style = bold_style

    workbook.save(filename)

    print(f"El reporte ha sido exportado como {filename}.")


# Creamos la opción Clientes
def Clientes():
    while True:
        mostrar_menu_clientes()
        opcion = input("Seleccione una opción: ")
        if opcion == "1":
            agregar_cliente()
        elif opcion == "2":
            listado_clientes()
        elif opcion == "3":
            print("Volviendo al menú principal.")
            break
        else:
            print("Opción no válida. Por favor, seleccione una opción válida.")

def agregar_cliente():
    # Solicitar el nombre del servicio al usuario y validar que no esté en blanco
    while True:
        nombre = input("Ingrese el nombre completo del cliente: ")
        if nombre.strip():  # Verifica que el nombre no esté en blanco
            break
        else:
            print("El nombre no puede estar en blanco. Por favor, ingréselo nuevamente.") 

    while True:
        rfc = input("Ingrese el RFC del cliente: ")
        if validar_rfc(rfc):
            print("RFC válido.")
            break  # Sale del bucle si el RFC es válido
        else:
            print("Error: el RFC no tiene un formato válido. Intente de nuevo.")

    while True:
        correo = input("Ingrese la dirección de correo electrónico: ")
        if validar_correo(correo):
            print("Correo electrónico válido.")
            break # Sale del bucle si el correo es válido
        else:
            print("Error: la dirección de correo electrónico no tiene un formato válido.")

    # Realizar la inserción en la base de datos
    with sqlite3.connect("taller.db") as conn:
        cursor = conn.cursor()
        try:
            cursor.execute("INSERT INTO Clientes (nombrecliente, rfccliente, correocliente) VALUES (?, ?, ?)", (nombre, rfc, correo))
            conn.commit()
            cursor.execute("SELECT idcliente FROM Clientes WHERE idcliente = (SELECT MAX(idcliente) FROM Clientes)")
            clavecliente = cursor.fetchone()
            print(f"Cliente agregado exitosamente con clave {clavecliente[0]}")
        except Exception as e:
            print(f"Error al agregar el servicio: {e}")


def listado_clientes():
    while True:
        print("\nListado de clientes registrados:")
        print("1. Ordenado por clave")
        print("2. Ordenado por nombre")
        print("3. Volver al menú anterior")
        opcion = input("Seleccione una opción: ")

        if opcion == "1":  # Listado de clientes ordenado por clave
            ordenar_clientes_por_clave()
        elif opcion == "2":  # Listado de clientes ordenado por nombre
            ordenar_clientes_por_nombre()
        elif opcion == "3":  # Volver al menú de consultas y reportes de clientes
            print("Volviendo al menú de consultas y reportes de clientes.")
            break
        else:
            print("Opción no válida. Por favor, seleccione una opción válida.")

def ordenar_clientes_por_clave():
    with sqlite3.connect("taller.db") as conn:
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT idcliente, nombrecliente, rfccliente, correocliente FROM Clientes ORDER BY idcliente")
            clientes = cursor.fetchall()
            if clientes:
                print("\nListado de clientes activos ordenado por clave:")
                for cliente in clientes:
                    print(f"Clave: {cliente[0]}, Nombre: {cliente[1]}, RFC: {cliente[2]}, Correo: {cliente[3]}")

                # Preguntar al usuario si desea exportar el reporte
                while True:
                    opcion_exportar = input("\n¿Desea exportar el reporte a CSV, Excel o regresar al menú de reportes? (CSV/Excel/Regresar): ").strip().lower()
                    if opcion_exportar == "csv":
                        exportar_reporte_clientes_csv(clientes, "ClientesActivosPorClave")
                        break
                    elif opcion_exportar == "excel":
                        exportar_reporte_clientes_excel(clientes, "ClientesActivosPorClave")
                        break
                    elif opcion_exportar == "regresar":
                        break
                    else:
                        print("Opción no válida. Por favor, seleccione una opción válida.")
            else:
                print("No hay clientes activos registrados.")
        except sqlite3.Error as e:
            print(f"Error al obtener el listado de clientes: {e}")

def exportar_reporte_clientes_csv(data, report_name):
    filename = f"{report_name}_{obtener_fecha_actual()}.csv"
    with open(filename, mode='w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(["Clave", "Nombre", "RFC", "Correo"])
        for cliente in data:
            writer.writerow(cliente)
    print(f"El reporte ha sido exportado como {filename}")

def exportar_reporte_clientes_excel(data, report_name):
    filename = f"{report_name}_{obtener_fecha_actual()}.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = report_name

    bold = openpyxl.styles.Font(bold=True)
    bold_style = openpyxl.styles.NamedStyle(name="bold")
    bold_style.font = bold

    worksheet['A1'] = "Clave"
    worksheet['B1'] = "Nombre"
    worksheet['C1'] = "RFC"
    worksheet['D1'] = "Correo"

    worksheet['A1'].style = bold_style
    worksheet['B1'].style = bold_style
    worksheet['C1'].style = bold_style
    worksheet['D1'].style = bold_style

    row = 2

    for cliente in data:
        worksheet.cell(row=row, column=1, value=cliente[0])
        worksheet.cell(row=row, column=2, value=cliente[1])
        worksheet.cell(row=row, column=3, value=cliente[2])
        worksheet.cell(row=row, column=4, value=cliente[3])
        row += 1

    workbook.save(filename)

    print(f"El reporte ha sido exportado como {filename}")

def obtener_fecha_actual():
    fecha_actual = datetime.datetime.now()
    return fecha_actual.strftime("%m_%d_%Y")

def ordenar_clientes_por_nombre():
    with sqlite3.connect("taller.db") as conn:
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT idcliente, nombrecliente, rfccliente, correocliente FROM Clientes ORDER BY UPPER(nombrecliente)")
            clientes = cursor.fetchall()
            if clientes:
                print("\nListado de clientes ordenado por nombre:")
                for cliente in clientes:
                    print(f"Clave: {cliente[0]}, Nombre: {cliente[1]}, RFC: {cliente[2]}, Correo: {cliente[3]}")

                # Preguntar al usuario si desea exportar el reporte
                while True:
                    opcion_exportar = input("\n¿Desea exportar el reporte a CSV, Excel o regresar al menú de reportes? (CSV/Excel/Regresar): ").strip().lower()
                    if opcion_exportar == "csv":
                        exportar_reporte_clientes_csv(clientes, "ClientesPorNombre")
                        break
                    elif opcion_exportar == "excel":
                        exportar_reporte_clientes_excel(clientes, "ClientesPorNombre")
                        break
                    elif opcion_exportar == "regresar":
                        break
                    else:
                        print("Opción no válida. Por favor, seleccione una opción válida.")
            else:
                print("No hay clientes registrados.")
        except sqlite3.Error as e:
            print(f"Error al obtener el listado de clientes: {e}")




# Creamos la opción Servicios
def Servicios():
    while True:
        mostrar_menu_servicios()
        opcion = input("Seleccione una opción: ")
        if opcion == "1":
            agregar_servicio()
        elif opcion == "2":
            consultasreportes_servicios()
        elif opcion == "3":
            print("Volviendo al menú principal.")
            break
        else:
            print("Opción no válida. Por favor, seleccione una opción válida.")

def agregar_servicio():
    # Solicitar el nombre del servicio al usuario y validar que no esté en blanco
    while True:
        nombre = input("Ingrese el nombre del servicio: ")
        if nombre.strip():  # Verifica que el nombre no esté en blanco
            break
        else:
            print("El nombre no puede estar en blanco. Por favor, ingréselo nuevamente.")
    
    try:
        costo = float(input("Ingrese el costo del servicio: "))
    except ValueError:
        print("El costo debe ser un número válido.")
        return
    # Realizar la inserción en la base de datos
    with sqlite3.connect("taller.db") as conn:
        cursor = conn.cursor()
        try:
            cursor.execute("INSERT INTO Servicios (nombreservicio, costoservicio) VALUES (?, ?)", (nombre, costo))
            conn.commit()
            cursor.execute("SELECT idservicio FROM Servicios WHERE idservicio = (SELECT MAX(idservicio) FROM Servicios)")
            claveservicio = cursor.fetchone()
            print(f"Servicio agregado exitosamente con clave {claveservicio[0]}")
        except Exception as e:
            print(f"Error al agregar el servicio: {e}")
def consultasreportes_servicios():
    while True:
        print("1. Búsqueda por clave de servicio")
        print("2. Búsqueda por nombre de servicio")
        print("3. Listado de servicios")
        print("4. Salir al menú Servicios")
        opcion = input("Seleccione una opción: ")
        if opcion == "1":    # Búsqueda por clave de servicio
            Busqueda_clave_servicio()
        elif opcion == "2":  # Búsqueda por nombre de servicio
            Busqueda_nombre_servicio()
        elif opcion == "3":  # Listado de servicios
            Listado_servicios()
        elif opcion == "4":  # Salir al menú Servicios
            print("Saliendo del programa. ¡Hasta luego!")
            break
        else:
            print("Opción no válida. Por favor, seleccione una opción válida.")

def Busqueda_clave_servicio():
    # Conectarse a la base de datos
    with sqlite3.connect("taller.db") as conn:
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT idservicio, nombreservicio FROM Servicios")
            servicios = cursor.fetchall()

            if not servicios:
                print("No hay servicios registrados en la base de datos.")
                return

            # Mostrar un listado tabular de claves y nombres de servicio
            print("Listado de servicios:")
            print("Clave  |  Nombre del Servicio")
            for servicio in servicios:
                print(f"{servicio[0]:<6} | {servicio[1]}")

            # Solicitar al usuario que ingrese la clave del servicio a consultar
            while True:
                try:
                    clave_servicio = int(input("Ingrese la clave del servicio que desea consultar: "))
                    if clave_servicio in [servicio[0] for servicio in servicios]:
                        break
                    else:
                        print("Clave no válida. Por favor, ingrese una clave válida.")
                except ValueError:
                    print("Error: La clave debe ser un número válido.")

            # Recuperar los detalles del servicio seleccionado
            cursor.execute("SELECT idservicio, nombreservicio, costoservicio FROM Servicios WHERE idservicio = ?", (clave_servicio,))
            servicio_detalle = cursor.fetchone()

            # Mostrar los detalles del servicio
            print("\nDetalle del servicio:")
            print(f"Clave del Servicio: {servicio_detalle[0]}")
            print(f"Nombre del Servicio: {servicio_detalle[1]}")
            print(f"Costo del Servicio: {servicio_detalle[2]:.2f}")

        except sqlite3.Error as e:
            print(f"Error al buscar el servicio: {e}")

def Busqueda_nombre_servicio():
    # Conectarse a la base de datos
    with sqlite3.connect("taller.db") as conn:
        cursor = conn.cursor()
        try:
            nombre_servicio_buscar = input("Ingrese el nombre del servicio a buscar: ").strip().lower()

            # Realizar una consulta insensible a mayúsculas y minúsculas
            cursor.execute("SELECT idservicio, nombreservicio, costoservicio FROM Servicios WHERE LOWER(nombreservicio) = ?", (nombre_servicio_buscar,))
            servicio_detalle = cursor.fetchone()

            if servicio_detalle:
                # Mostrar los detalles del servicio encontrado
                print("\nDetalle del servicio encontrado:")
                print(f"Clave del Servicio: {servicio_detalle[0]}")
                print(f"Nombre del Servicio: {servicio_detalle[1]}")
                print(f"Costo del Servicio: {servicio_detalle[2]:.2f}")
            else:
                print(f"No se encontró un servicio con el nombre '{nombre_servicio_buscar}'.")

        except sqlite3.Error as e:
            print(f"Error al buscar el servicio: {e}")

def Listado_servicios():
    while True:
        print("\nListado de servicios:")
        print("1. Ordenado por clave")
        print("2. Ordenado por nombre de servicio")
        print("3. Regresar al menú anterior")
        opcion = input("Seleccione una opción: ")

        if opcion == "1":  # Listado de servicios ordenado por clave
            ordenar_servicios_por_clave()
        elif opcion == "2":  # Listado de servicios ordenado por nombre
            ordenar_servicios_por_nombre()
        elif opcion == "3":  # Volver al menú de consultas y reportes de servicios
            print("Volviendo al menú de consultas y reportes de servicios.")
            break
        else:
            print("Opción no válida. Por favor, seleccione una opción válida.")

def ordenar_servicios_por_clave():
    with sqlite3.connect("taller.db") as conn:
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT idservicio, nombreservicio, costoservicio FROM Servicios ORDER BY idservicio")
            servicios = cursor.fetchall()
            if servicios:
                print("\nListado de servicios ordenado por clave:")
                for servicio in servicios:
                    print(f"Clave: {servicio[0]}, Nombre: {servicio[1]}, Costo: {servicio[2]:.2f}")

                # Preguntar al usuario si desea exportar el reporte
                exportar_reporte_servicios(servicios, "ServiciosPorClave")

            else:
                print("No hay servicios registrados.")
        except sqlite3.Error as e:
            print(f"Error al obtener el listado de servicios: {e}")

def ordenar_servicios_por_nombre():
    with sqlite3.connect("taller.db") as conn:
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT idservicio, nombreservicio, costoservicio FROM Servicios ORDER BY UPPER(nombreservicio)")
            servicios = cursor.fetchall()
            if servicios:
                print("\nListado de servicios ordenado por nombre:")
                for servicio in servicios:
                    print(f"Clave: {servicio[0]}, Nombre: {servicio[1]}, Costo: {servicio[2]:.2f}")

                # Preguntar al usuario si desea exportar el reporte
                exportar_reporte_servicios(servicios, "ServiciosPorNombre")
            else:
                print("No hay servicios registrados.")
        except sqlite3.Error as e:
            print(f"Error al obtener el listado de servicios: {e}")

def exportar_reporte_servicios(data, report_name):
    while True:
        opcion_exportar = input("\n¿Desea exportar el reporte a CSV, Excel o regresar al menú de reportes? (CSV/Excel/Regresar): ").strip().lower()
        if opcion_exportar == "csv":
            exportar_reporte_servicios_csv(data, report_name)
            break
        elif opcion_exportar == "excel":
            exportar_reporte_servicios_excel(data, report_name)
            break
        elif opcion_exportar == "regresar":
            break
        else:
            print("Opción no válida. Por favor, seleccione una opción válida.")

def exportar_reporte_servicios_csv(data, report_name):
    filename = f"{report_name}_{obtener_fecha_actual()}.csv"
    with open(filename, mode='w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(["Clave", "Nombre", "Costo"])
        for servicio in data:
            writer.writerow(servicio)
    print(f"El reporte ha sido exportado como {filename}")

def exportar_reporte_servicios_excel(data, report_name):
    filename = f"{report_name}_{obtener_fecha_actual()}.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = report_name

    bold = openpyxl.styles.Font(bold=True)
    bold_style = openpyxl.styles.NamedStyle(name="bold")
    bold_style.font = bold

    worksheet['A1'] = "Clave"
    worksheet['B1'] = "Nombre"
    worksheet['C1'] = "Costo"

    worksheet['A1'].style = bold_style
    worksheet['B1'].style = bold_style
    worksheet['C1'].style = bold_style

    row = 2

    for servicio in data:
        worksheet.cell(row=row, column=1, value=servicio[0])
        worksheet.cell(row=row, column=2, value=servicio[1])
        worksheet.cell(row=row, column=3, value=servicio[2])
        row += 1

    workbook.save(filename)
    print(f"El reporte ha sido exportado como {filename}")

def obtener_fecha_actual():
    return datetime.now()

def validar_fecha(fecha):
    try:
        fecha_ingresada = datetime.strptime(fecha, '%d/%m/%Y')
        fecha_actual = obtener_fecha_actual()
        if fecha_ingresada <= fecha_actual:
            return True
        else:
            return False
    except ValueError:
        return False

# Función para generar una clave aleatoria de 6 caracteres compuesta por letras y dígitos
def generar_clave_aleatoria():
    caracteres = string.ascii_uppercase + string.digits
    return ''.join(random.choice(caracteres) for _ in range(6))

# Función para validar que el rfc tenga el formato correcto
def validar_rfc(rfc):
    # Patrones de expresiones regulares para RFC de personas físicas y morales
    rfc_fisica_pattern = r'^[A-Z]{4}\d{2}(0[1-9]|1[0-2])(0[1-9]|[1-2]\d|3[0-1])[A-Z\d]{3}$'
    rfc_moral_pattern = r'^[A-Z]{3}\d{2}(0[1-9]|1[0-2])(0[1-9]|[1-2]\d|3[0-1])[A-Z\d]{3}$'
    if re.match(rfc_fisica_pattern, rfc) or re.match(rfc_moral_pattern, rfc):
        return True
    else:
        return False

# Función para validar que el correo tenga el formato correcto
def validar_correo(correo):
    # Patrón de expresión regular para validar un correo electrónico
    correo_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    if re.match(correo_pattern, correo):
        return True
    else:
        return False
    
# Función para comprobar si una clave existe en la tabla especificada
def clavecliente_existe_(tabla, clave):
    with sqlite3.connect("taller.db") as conn:
        cursor = conn.cursor()
        try:
            cursor.execute(f"SELECT 1 FROM {tabla} WHERE idcliente = ?", (clave))
            return True
        except sqlite3.Error:
            return False

        
def claveservicio_existe_(tabla, clave):
    with sqlite3.connect("taller.db") as conn:
        cursor = conn.cursor()
        try:
            cursor.execute(f"SELECT 1 FROM {tabla} WHERE {tabla}idservicio = ?", (clave,))
            return cursor.fetchone() is not None
        except sqlite3.Error:
            return False
        
def mostrar_registros(tabla):
    print(f"\nRegistros de {tabla}")
    with sqlite3.connect("taller.db") as conn:
        cursor = conn.cursor()
        try:
            cursor.execute(f"SELECT * FROM {tabla}")
            rows = cursor.fetchall()
            for row in rows:
                print(row)
        except sqlite3.Error as e:
            print(f"Error al mostrar los registros de la tabla {tabla}: {e}")

# Función main que se ejecuta al iniciar el programa después de crear las tablas
def main():
    while True:
        mostrar_menu_principal()
        opcion = input("Seleccione una opción: ")
        if opcion == "1":    # Opción para submenú Notas
            Notas()
        elif opcion == "2":  # Opción para submenú Clientes
            Clientes()
        elif opcion == "3":  # Opción para submenú Servicios
            Servicios()
        elif opcion == "4":  # Opción para salir del programa
            confirmacion = input("¿Está seguro de que desea salir del programa? (Sí/No): ").strip().lower()
            if confirmacion == "si" or confirmacion == "sí":
                print("Saliendo del programa. ¡Hasta luego!")
                break
            elif confirmacion == "no":
                print("Regresando al menú principal.")
            else:
                print("Respuesta no válida. Por favor, seleccione 'Sí' o 'No'.")
        else:
            print("Opción no válida. Por favor, seleccione una opción válida.")

def mostrar_menu_principal():
    print("\nMenú Principal:")
    print("1. Notas")
    print("2. Clientes")
    print("3. Servicios")
    print("4. Salir")

if __name__ == "__main__":
    crear_tablas()
    main()
