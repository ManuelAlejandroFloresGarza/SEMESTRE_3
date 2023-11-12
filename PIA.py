import sqlite3
import random
import string
import re
import datetime
from datetime import datetime
import csv
import openpyxl
import os
from collections import Counter

def crear_tablas():
    try:
        with sqlite3.connect("taller.db") as conn:
            cursor = conn.cursor()
            cursor.execute('''CREATE TABLE IF NOT EXISTS Clientes (
                    idcliente INTEGER PRIMARY KEY AUTOINCREMENT,
                    nombrecliente TEXT NOT NULL,
                    rfccliente TEXT NOT NULL,
                    correocliente TEXT,
                    status INTEGER DEFAULT 1
                )''')

            cursor.execute('''CREATE TABLE IF NOT EXISTS Servicios (
                    idservicio INTEGER PRIMARY KEY AUTOINCREMENT,
                    nombreservicio TEXT NOT NULL,
                    costoservicio REAL NOT NULL,
                    status INTEGER DEFAULT 1
                )''')

            cursor.execute("""
            CREATE TABLE IF NOT EXISTS Notas (
                folionota INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha TEXT,
                idcliente INTEGER NOT NULL,
                idservicio INTEGER NOT NULL,
                status INTEGER DEFAULT 1,
                FOREIGN KEY (idcliente) REFERENCES Clientes(idcliente),
                FOREIGN KEY (idservicio) REFERENCES Servicios(idservicio)
            )
            """)
    except sqlite3.Error as e:
        print(f"Error al crear tablas en la base de datos: {e}")

def mostrar_menu_notas():
    print("\nMenú Notas:")
    print("1. Registrar una nota")
    print("2. Cancelar una nota")
    print("3. Recuperar una nota")
    print("4. Consultas y reportes de notas")
    print("5. Volver al menú principal")

def mostrar_menu_clientes():
    print("\nMenú de Clientes:")
    print("1. Agregar Cliente")
    print("2. Consultas y Reportes")
    print("3. Cancelar Cliente")
    print("4. Recuperar Cliente")
    print("5. Regresar al Menú Principal")

def mostrar_menu_servicios():
    print("\nMenú de Servicios:")
    print("1. Agregar Servicio")
    print("2. Consultas y Reportes")
    print("3. Cancelar Servicio")
    print("4. Recuperar Servicio")
    print("5. Regresar al Menú Principal")

def mostrar_menu_estadisticas():
    print("\nMenú Estadisticas:")
    print("1. Servicios más prestados")
    print("2. Clientes con más notas")
    print("3. Promedio de los montos de las notas")
    print("4. Volver al menú principal")

def Notas():
    while True:
        mostrar_menu_notas()
        opcion = input("Seleccione una opción: ")

        if opcion == "1":
            registrar_nota()
        elif opcion == "2":
            cancelar_nota()
        elif opcion == "3":
            recuperar_nota()
        elif opcion == "4":
            consultasreportes_notas()
        elif opcion == "5":
            print("Volviendo al menú principal.")
            break
        else:
            print("Opción no válida. Por favor, seleccione una opción válida.")
            
def registrar_nota():
    while True:
        while True:
            fecha_ingresada = input("Ingresa una fecha en el formato DD/MM/YYYY: ")
            if validar_fecha(fecha_ingresada):
                break
            else:
                print("Fecha no válida o es posterior a la fecha actual. Intenta de nuevo.")

        with sqlite3.connect("taller.db") as conn:
            cursor = conn.cursor()
            mostrar_registros("Clientes")
            mostrar_registros("Servicios")

            while True:
                try:
                    cliente = input("\nIngrese la clave del cliente: ")
                    cliente = int(cliente)
                    cursor.execute("SELECT idcliente, nombrecliente FROM Clientes WHERE idcliente = ?", (cliente,))
                    cliente_existente = cursor.fetchone()
                    if cliente_existente:
                        print(f"Cliente seleccionado: {cliente_existente[1]}")
                        break
                    else:
                        print("Clave de cliente no válida. Debe ser mayor que 0 y existir en la tabla Clientes. Intente de nuevo.")
                except ValueError:
                    print("Error: Ingrese una clave de cliente válida.")

            while True:
                try:
                    servicio = int(input("Ingrese la clave del servicio: "))
                    cursor.execute("SELECT idservicio, nombreservicio FROM Servicios WHERE idservicio = ?", (servicio,))
                    servicio_existente = cursor.fetchone()
                    if servicio_existente:
                        print(f"Servicio seleccionado: {servicio_existente[1]}")
                        break
                    else:
                        print("Clave de servicio no válida. Debe ser mayor que 0 y existir en la tabla Servicios. Intente de nuevo.")
                except ValueError:
                    print("Error: Ingrese una clave de servicio válida.")

            try:
                cursor.execute("INSERT INTO Notas (fecha, idcliente, idservicio) VALUES (?, ?, ?)", (fecha_ingresada, cliente, servicio))
                conn.commit()
                cursor.execute("SELECT folionota FROM Notas WHERE folionota = (SELECT MAX(folionota) FROM Notas)")
                folio = cursor.fetchone()
                print(f"Nota agregada exitosamente con folio {folio[0]}")
                break
            except Exception as e:
                print(f"Error al agregar la nota: {e}")

def cancelar_nota():
    while True:
        try:
            folio = int(input("Ingrese el folio de la nota que desea cancelar: "))
        except ValueError:
            print("Error: El folio debe ser un número válido.")
            continue

        with sqlite3.connect("taller.db") as conn:
            cursor = conn.cursor()

            cursor.execute("SELECT * FROM Notas WHERE folionota = ? AND status = 1", (folio,))
            nota = cursor.fetchone()

            if nota:
                print("Detalle de la nota a cancelar:")
                print(f"Folio de la nota: {nota[0]}")
                print(f"Fecha: {nota[1]}")
                cursor.execute("SELECT nombrecliente FROM Clientes WHERE idcliente = ?", (nota[2],))
                nombre_cliente = cursor.fetchone()[0]
                cursor.execute("SELECT nombreservicio FROM Servicios WHERE idservicio = ?", (nota[3],))
                nombre_servicio = cursor.fetchone()[0]
                print(f"Cliente: {nombre_cliente}")
                print(f"Servicio: {nombre_servicio}")

                confirmacion = input("¿Está seguro de que desea cancelar esta nota? (S/N): ").strip().upper()

                if confirmacion == 'S':
                    cursor.execute("UPDATE Notas SET status = 0 WHERE folionota = ?", (folio,))
                    conn.commit()
                    print("La nota ha sido cancelada exitosamente.")
                else:
                    print("La cancelación de la nota ha sido cancelada a solicitud del usuario.")
            else:
                print("La nota no existe o ya está cancelada en el sistema.")

        break


def recuperar_nota():
    with sqlite3.connect("taller.db") as conn:
        cursor = conn.cursor()

        cursor.execute("SELECT folionota FROM Notas WHERE status = 0")
        notas_canceladas = cursor.fetchall()

        if not notas_canceladas:
            print("No hay notas canceladas para recuperar.")
            return

        print("Listado de notas canceladas:")
        for folio in notas_canceladas:
            print(f"Folio de la nota: {folio[0]}")

        while True:
            try:
                folio_recuperar = int(input("Ingrese el folio de la nota que desea recuperar o 0 para cancelar: "))
            except ValueError:
                print("Error: El folio debe ser un número válido.")
                continue

            if folio_recuperar == 0:
                print("Operación cancelada. No se recuperó ninguna nota.")
                return
            elif (folio_recuperar,) in notas_canceladas:
                cursor.execute("SELECT * FROM Notas WHERE folionota = ?", (folio_recuperar,))
                nota = cursor.fetchone()

                print("Detalle de la nota a recuperar:")
                print(f"Folio de la nota: {nota[0]}")
                print(f"Fecha: {nota[1]}")
                cursor.execute("SELECT nombrecliente FROM Clientes WHERE idcliente = ?", (nota[2],))
                nombre_cliente = cursor.fetchone()[0]
                cursor.execute("SELECT nombreservicio FROM Servicios WHERE idservicio = ?", (nota[3],))
                nombre_servicio = cursor.fetchone()[0]
                print(f"Cliente: {nombre_cliente}")
                print(f"Servicio: {nombre_servicio}")

                confirmacion = input("¿Está seguro de que desea recuperar esta nota? (S/N): ").strip().upper()

                if confirmacion == 'S':
                    cursor.execute("UPDATE Notas SET status = 1 WHERE folionota = ?", (folio_recuperar,))
                    conn.commit()
                    print("La nota ha sido recuperada exitosamente.")
                else:
                    print("La recuperación de la nota ha sido cancelada a solicitud del usuario.")
                break
            else:
                print("El folio ingresado no corresponde a una nota cancelada en el sistema.")

def consultasreportes_notas():
    while True:
        print("\nConsultas y Reportes de Notas:")
        print("1. Consulta por período")
        print("2. Consulta por folio")
        print("3. Volver atrás")
        opcion = input("Seleccione una opción: ")

        if opcion == "1":  
            consulta_por_periodo()
        elif opcion == "2":  
            consulta_por_folio()
        elif opcion == "3":  
            print("Volviendo al menú principal.")
            break
        else:
            print("Opción no válida. Por favor, seleccione una opción válida.")

def consulta_por_periodo():
    while True:
        fecha_inicial = input("Ingrese la fecha inicial (DD-MM-YYYY) o presione Enter para asumir 01/01/2000: ")
        print(fecha_inicial)
        fecha_final = input("Ingrese la fecha final (DD-MM-YYYY) o presione Enter para asumir la fecha actual: ")
        print(fecha_final)

        if fecha_inicial == "":
            fecha_inicial = "01-01-2000"
        if fecha_final == "":
            fecha_final = datetime.now().strftime("%d-%m-%Y")

        try:
            fecha_inicial = datetime.strptime(fecha_inicial, '%d-%m-%Y')
            fecha_final = datetime.strptime(fecha_final, '%d-%m-%Y')
        except ValueError:
            print("Error: Las fechas no tienen un formato válido (DD-MM-YYYY).")
            continue

        if fecha_final < fecha_inicial:
            print("Error: La fecha final debe ser igual o posterior a la fecha inicial.")
            continue

        with sqlite3.connect("taller.db") as conn:
            cursor = conn.cursor()

            cursor.execute("""
                SELECT n.folionota, n.fecha, c.nombrecliente
                FROM Notas n
                JOIN Clientes c ON n.idcliente = c.idcliente
                WHERE date(n.fecha) BETWEEN ? AND ?
            """, (fecha_inicial, fecha_final))
            notas = cursor.fetchall()

            if notas:
                cursor.execute("""
                    SELECT AVG(s.costoservicio)
                    FROM Servicios s
                    JOIN Notas n ON s.idservicio = n.idservicio
                    WHERE date(n.fecha) BETWEEN ? AND ?
                """, (fecha_inicial, fecha_final))
                promedio_monto = cursor.fetchone()[0]

                print("Notas dentro del período seleccionado:")
                print(f"{'Folio':<10} {'Fecha':<12} {'Nombre del Cliente':<30}")
                for nota in notas:
                    print(f"{nota[0]:<10} {nota[1]:<12} {nota[2]:<30}")

                print(f"Monto promedio del período: {promedio_monto:.2f}")

                while True:
                    opcion_exportar = input("¿Desea exportar el reporte a CSV, Excel o regresar al menú de reportes? (CSV/Excel/Regresar): ").strip().lower()
                    if opcion_exportar == "csv":
                        exportar_reporte_csv(notas, fecha_inicial, fecha_final, promedio_monto)
                        break
                    elif opcion_exportar == "excel":
                        exportar_reporte_excel(notas, fecha_inicial, fecha_final, promedio_monto)
                        break
                    elif opcion_exportar == "regresar":
                        break
                    else:
                        print("Opción no válida. Por favor, seleccione una opción válida.")
            else:
                print("No hay notas emitidas para el período especificado.")
            break

def consulta_por_folio():
    while True:
        with sqlite3.connect("taller.db") as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT n.folionota, n.fecha, c.nombrecliente, s.nombreservicio
                FROM Notas n
                JOIN Clientes c ON n.idcliente = c.idcliente
                JOIN Servicios s ON n.idservicio = s.idservicio
                WHERE n.status = 1
                ORDER BY n.folionota
            """)
            notas = cursor.fetchall()

            if not notas:
                print("No hay notas activas en el sistema para consultar.")
                return

            print("Listado de notas activas:")
            print("Folio  |  Fecha       |  Nombre del Cliente  |  Servicio")
            for nota in notas:
                print(f"{nota[0]:<7} | {nota[1]} | {nota[2]} | {nota[3]}")

            try:
                folio_consultar = int(input("Ingrese el folio de la nota que desea consultar o 0 para salir: "))
            except ValueError:
                print("Error: El folio debe ser un número válido.")
                continue

            if folio_consultar == 0:
                print("Saliendo del menú de consulta por folio.")
                break

            if (folio_consultar,) in [(n[0],) for n in notas]:
                cursor.execute("""
                    SELECT n.folionota, n.fecha, c.nombrecliente, c.rfccliente, c.correocliente, s.nombreservicio, s.costoservicio
                    FROM Notas n
                    JOIN Clientes c ON n.idcliente = c.idcliente
                    JOIN Servicios s ON n.idservicio = s.idservicio
                    WHERE n.folionota = ? AND n.status = 1
                """, (folio_consultar,))
                nota = cursor.fetchone()

                print("Detalle de la nota:")
                print(f"Folio de la nota: {nota[0]}")
                print(f"Fecha: {nota[1]}")
                print(f"Nombre del Cliente: {nota[2]}")
                print(f"RFC del Cliente: {nota[3]}")
                print(f"Correo del Cliente: {nota[4]}")
                print(f"Servicio: {nota[5]}")
                print(f"Costo del servicio: {nota[6]:.2f}")
            else:
                print("El folio ingresado no corresponde a una nota activa en el sistema o no existe.")




def exportar_reporte_csv(notas, fecha_inicial, fecha_final, promedio_monto):
    if fecha_inicial is not None and fecha_final is not None:
        filename = f"ReportePorPeriodo_{fecha_inicial.strftime('%m_%d_%Y')}_{fecha_final.strftime('%m_%d_%Y')}.csv"
    else:
        filename = "ReporteNotas.csv"

    with open(filename, mode='w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(["Folio", "Fecha", "Nombre del Cliente"])
        for nota in notas:
            writer.writerow(nota)
        if fecha_inicial is not None and fecha_final is not None:
            writer.writerow(["Monto promedio del período:", "", promedio_monto])

    print(f"El reporte ha sido exportado como {filename}.")

def exportar_reporte_excel(notas, fecha_inicial, fecha_final, promedio_monto):
    if fecha_inicial is not None and fecha_final is not None:
        filename = f"ReportePorPeriodo_{fecha_inicial.strftime('%m_%d_%Y')}_{fecha_final.strftime('%m_%d_%Y')}.xlsx"
    else:
        filename = "ReporteNotas.xlsx"

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Reporte de Notas"

    bold = openpyxl.styles.Font(bold=True)
    bold_style = openpyxl.styles.NamedStyle(name="bold")
    bold_style.font = bold

    worksheet['A1'] = "Folio"
    worksheet['B1'] = "Fecha"
    worksheet['C1'] = "Nombre del Cliente"

    worksheet['A1'].style = bold_style
    worksheet['B1'].style = bold_style
    worksheet['C1'].style = bold_style

    row = 2

    for nota in notas:
        worksheet.cell(row=row, column=1, value=nota[0])
        worksheet.cell(row=row, column=2, value=nota[1])
        worksheet.cell(row=row, column=3, value=nota[2])
        row += 1

    if fecha_inicial is not None and fecha_final is not None:
        worksheet.cell(row=row, column=1, value="Monto promedio del período:")
        worksheet.cell(row=row, column=3, value=promedio_monto)
        worksheet.cell(row=row, column=1).style = bold_style
        worksheet.cell(row=row, column=3).style = bold_style

    workbook.save(filename)

    print(f"El reporte ha sido exportado como {filename}.")


def Clientes():
    while True:
        mostrar_menu_clientes()
        opcion = input("Seleccione una opción: ")
        if opcion == "1":
            agregar_cliente()
        elif opcion == "2":
            listado_clientes()
        elif opcion == "3":
            cancelar_cliente()
        elif opcion == "4":
            recuperar_cliente()
        elif opcion == "5":
            print("Volviendo al menú principal.")
            break
        else:
            print("Opción no válida. Por favor, seleccione una opción válida.")

def agregar_cliente():

    while True:

        nombre = input("Ingrese el nombre completo del cliente: ")

        if nombre.strip(): 

            break

        else:

            print("El nombre no puede estar en blanco. Por favor, ingréselo nuevamente.") 

    while True:

        rfc = input("Ingrese el RFC del cliente: ")

        if validar_rfc(rfc):

            print("RFC válido.")

            break  

        else:

            print("Error: el RFC no tiene un formato válido. Intente de nuevo.")

    while True:

        correo = input("Ingrese la dirección de correo electrónico: ")

        if validar_correo(correo):

            print("Correo electrónico válido.")

            break 

        else:

            print("Error: la dirección de correo electrónico no tiene un formato válido.")

    with sqlite3.connect("taller.db") as conn:

        cursor = conn.cursor()

        try:

            cursor.execute("INSERT INTO Clientes (nombrecliente, rfccliente, correocliente) VALUES (?, ?, ?)", (nombre, rfc, correo))

            conn.commit()

            cursor.execute("SELECT idcliente FROM Clientes WHERE idcliente = (SELECT MAX(idcliente) FROM Clientes)")

            clavecliente = cursor.fetchone()

            print(f"Cliente agregado exitosamente con clave {clavecliente[0]}")

        except Exception as e:

            print(f"Error al agregar el servicio: {e}")

def cancelar_cliente():

    while True:

        try:

            clave_cliente = int(input("Ingrese la clave del cliente que desea cancelar: "))

        except ValueError:

            print("Error: La clave del cliente debe ser un número válido.")

            continue



        with sqlite3.connect("taller.db") as conn:

            cursor = conn.cursor()



            cursor.execute("SELECT * FROM Clientes WHERE idcliente = ? AND status = 1", (clave_cliente,))

            cliente = cursor.fetchone()



            if cliente:

                print("Detalle del cliente a cancelar:")

                print(f"Clave del cliente: {cliente[0]}")

                print(f"Nombre: {cliente[1]}")

                print(f"RFC: {cliente[2]}")

                print(f"Correo: {cliente[3]}")



                confirmacion = input("¿Está seguro de que desea cancelar este cliente? (S/N): ").strip().upper()



                if confirmacion == 'S':

                    cursor.execute("UPDATE Clientes SET status = 0 WHERE idcliente = ?", (clave_cliente,))

                    conn.commit()

                    print("El cliente ha sido cancelado exitosamente.")

                else:

                    print("La cancelación del cliente ha sido cancelada a solicitud del usuario.")

            else:

                print("El cliente no existe o ya está cancelado en el sistema.")



        break



def recuperar_cliente():

    with sqlite3.connect("taller.db") as conn:

        cursor = conn.cursor()



        cursor.execute("SELECT idcliente, nombrecliente, rfccliente, correocliente FROM Clientes WHERE status = 0")

        clientes_cancelados = cursor.fetchall()



        if not clientes_cancelados:

            print("No hay clientes cancelados para recuperar.")

            return



        print("Listado de clientes cancelados:")

        print("Clave  |  Nombre  |  RFC  |  Correo")

        for cliente in clientes_cancelados:

            print(f"{cliente[0]:<6} | {cliente[1]} | {cliente[2]} | {cliente[3]}")



        while True:

            try:

                clave_recuperar = int(input("Ingrese la clave del cliente que desea recuperar o 0 para cancelar: "))

            except ValueError:

                print("Error: La clave del cliente debe ser un número válido.")

                continue



            if clave_recuperar == 0:

                print("Operación cancelada. No se recuperó ningún cliente.")

                return

            elif any(clave_recuperar == cliente[0] for cliente in clientes_cancelados):

                cursor.execute("SELECT * FROM Clientes WHERE idcliente = ?", (clave_recuperar,))

                cliente = cursor.fetchone()



                print("\nDetalle del cliente a recuperar:")

                print(f"Clave del cliente: {cliente[0]}")

                print(f"Nombre: {cliente[1]}")

                print(f"RFC: {cliente[2]}")

                print(f"Correo: {cliente[3]}")



                confirmacion = input("¿Está seguro de que desea recuperar este cliente? (S/N): ").strip().upper()



                if confirmacion == 'S':

                    cursor.execute("UPDATE Clientes SET status = 1 WHERE idcliente = ?", (clave_recuperar,))

                    conn.commit()

                    print("El cliente ha sido recuperado exitosamente.")

                else:

                    print("La recuperación del cliente ha sido cancelada a solicitud del usuario.")

                break

            else:

                print("La clave ingresada no corresponde a un cliente cancelado en el sistema.")

                break







def listado_clientes():

    while True:

        print("\nListado de clientes registrados:")

        print("1. Ordenado por clave")

        print("2. Ordenado por nombre")

        print("3. Volver al menú anterior")

        opcion = input("Seleccione una opción: ")



        if opcion == "1":  

            ordenar_clientes_por_clave()

        elif opcion == "2":  

            ordenar_clientes_por_nombre()

        elif opcion == "3": 

            print("Volviendo al menú de consultas y reportes de clientes.")

            break

        else:

            print("Opción no válida. Por favor, seleccione una opción válida.")



def ordenar_clientes_por_clave():

    with sqlite3.connect("taller.db") as conn:

        cursor = conn.cursor()

        try:

            cursor.execute("SELECT idcliente, nombrecliente, rfccliente, correocliente FROM Clientes WHERE status = 1 ORDER BY idcliente")

            clientes = cursor.fetchall()

            if clientes:

                print("\nListado de clientes activos ordenado por clave:")

                for cliente in clientes:

                    print(f"Clave: {cliente[0]}, Nombre: {cliente[1]}, RFC: {cliente[2]}, Correo: {cliente[3]}")



                while True:

                    opcion_exportar = input("\n¿Desea exportar el reporte a CSV, Excel o regresar al menú de reportes? (CSV/Excel/Regresar): ").strip().lower()

                    if opcion_exportar == "csv":

                        exportar_reporte_clientes_csv(clientes, "ClientesActivosPorClave")

                        break

                    elif opcion_exportar == "excel":

                        exportar_reporte_clientes_excel(clientes, "ClientesActivosPorClave")

                        break

                    elif opcion_exportar == "regresar":

                        break

                    else:

                        print("Opción no válida. Por favor, seleccione una opción válida.")

            else:

                print("No hay clientes activos registrados.")

        except sqlite3.Error as e:

            print(f"Error al obtener el listado de clientes: {e}")



def exportar_reporte_clientes_csv(data, report_name):

    filename = f"{report_name}_{obtener_fecha_actual()}.csv"

    with open(filename, mode='w', newline='') as file:

        writer = csv.writer(file)

        writer.writerow(["Clave", "Nombre", "RFC", "Correo"])

        for cliente in data:

            writer.writerow(cliente)

    print(f"El reporte ha sido exportado como {filename}")



def exportar_reporte_clientes_excel(data, report_name):

    filename = f"{report_name}_{obtener_fecha_actual()}.xlsx"

    workbook = openpyxl.Workbook()

    worksheet = workbook.active

    worksheet.title = report_name



    bold = openpyxl.styles.Font(bold=True)

    bold_style = openpyxl.styles.NamedStyle(name="bold")

    bold_style.font = bold



    worksheet['A1'] = "Clave"

    worksheet['B1'] = "Nombre"

    worksheet['C1'] = "RFC"

    worksheet['D1'] = "Correo"



    worksheet['A1'].style = bold_style

    worksheet['B1'].style = bold_style

    worksheet['C1'].style = bold_style

    worksheet['D1'].style = bold_style



    row = 2



    for cliente in data:

        worksheet.cell(row=row, column=1, value=cliente[0])

        worksheet.cell(row=row, column=2, value=cliente[1])

        worksheet.cell(row=row, column=3, value=cliente[2])

        worksheet.cell(row=row, column=4, value=cliente[3])

        row += 1



    workbook.save(filename)



    print(f"El reporte ha sido exportado como {filename}")



def obtener_fecha_actual():

    fecha_actual = datetime.datetime.now()

    return fecha_actual.strftime("%m_%d_%Y")



def ordenar_clientes_por_nombre():

    with sqlite3.connect("taller.db") as conn:

        cursor = conn.cursor()

        try:

            cursor.execute("SELECT idcliente, nombrecliente, rfccliente, correocliente FROM Clientes WHERE status=1 ORDER BY UPPER(nombrecliente)")

            clientes = cursor.fetchall()

            if clientes:

                print("\nListado de clientes ordenado por nombre:")

                for cliente in clientes:

                    print(f"Clave: {cliente[0]}, Nombre: {cliente[1]}, RFC: {cliente[2]}, Correo: {cliente[3]}")



                while True:

                    opcion_exportar = input("\n¿Desea exportar el reporte a CSV, Excel o regresar al menú de reportes? (CSV/Excel/Regresar): ").strip().lower()

                    if opcion_exportar == "csv":

                        exportar_reporte_clientes_csv(clientes, "ClientesPorNombre")

                        break

                    elif opcion_exportar == "excel":

                        exportar_reporte_clientes_excel(clientes, "ClientesPorNombre")

                        break

                    elif opcion_exportar == "regresar":

                        break

                    else:

                        print("Opción no válida. Por favor, seleccione una opción válida.")

            else:

                print("No hay clientes registrados.")

        except sqlite3.Error as e:

            print(f"Error al obtener el listado de clientes: {e}")









def Servicios():

    while True:

        mostrar_menu_servicios()

        opcion = input("Seleccione una opción: ")

        if opcion == "1":

            agregar_servicio()

        elif opcion == "2":

            consultasreportes_servicios()

        elif opcion == "3":

            cancelar_servicio()

        elif opcion == "4":

            recuperar_servicio()

        elif opcion == "5":

            print("Volviendo al menú principal.")

            break

        else:

            print("Opción no válida. Por favor, seleccione una opción válida.")





def agregar_servicio():

    while True:

        nombre = input("Ingrese el nombre del servicio: ")

        if nombre.strip():  

            break

        else:

            print("El nombre no puede estar en blanco. Por favor, ingréselo nuevamente.")

    

    try:

        costo = float(input("Ingrese el costo del servicio: "))

    except ValueError:

        print("El costo debe ser un número válido.")

        return

    with sqlite3.connect("taller.db") as conn:

        cursor = conn.cursor()

        try:

            cursor.execute("INSERT INTO Servicios (nombreservicio, costoservicio) VALUES (?, ?)", (nombre, costo))

            conn.commit()

            cursor.execute("SELECT idservicio FROM Servicios WHERE idservicio = (SELECT MAX(idservicio) FROM Servicios)")

            claveservicio = cursor.fetchone()

            print(f"Servicio agregado exitosamente con clave {claveservicio[0]}")

        except Exception as e:

            print(f"Error al agregar el servicio: {e}")



def cancelar_servicio():

    while True:

        try:

            clave_servicio = int(input("Ingrese la clave del servicio que desea cancelar: "))

        except ValueError:

            print("Error: La clave del servicio debe ser un número válido.")

            continue



        with sqlite3.connect("taller.db") as conn:

            cursor = conn.cursor()



            cursor.execute("SELECT * FROM Servicios WHERE idservicio = ? AND status = 1", (clave_servicio,))

            servicio = cursor.fetchone()



            if servicio:

                print("Detalle del servicio a cancelar:")

                print(f"Clave del servicio: {servicio[0]}")

                print(f"Nombre: {servicio[1]}")

                print(f"Costo: {servicio[2]:.2f}")



                confirmacion = input("¿Está seguro de que desea cancelar este servicio? (S/N): ").strip().upper()



                if confirmacion == 'S':

                    cursor.execute("UPDATE Servicios SET status = 0 WHERE idservicio = ?", (clave_servicio,))

                    conn.commit()

                    print("El servicio ha sido cancelado exitosamente.")

                else:

                    print("La cancelación del servicio ha sido cancelada a solicitud del usuario.")

            else:

                print("El servicio no existe o ya está cancelado en el sistema.")



        break



def recuperar_servicio():

    with sqlite3.connect("taller.db") as conn:

        cursor = conn.cursor()



        cursor.execute("SELECT idservicio, nombreservicio, costoservicio FROM Servicios WHERE status = 0")

        servicios_cancelados = cursor.fetchall()



        if not servicios_cancelados:

            print("No hay servicios cancelados para recuperar.")

            return



        print("Listado de servicios cancelados:")

        print("Clave  |  Nombre  |  Costo")

        for servicio in servicios_cancelados:

            print(f"{servicio[0]:<6} | {servicio[1]} | {servicio[2]:.2f}")



        while True:

            try:

                clave_recuperar = int(input("Ingrese la clave del servicio que desea recuperar o 0 para cancelar: "))

            except ValueError:

                print("Error: La clave del servicio debe ser un número válido.")

                continue



            if clave_recuperar == 0:

                print("Operación cancelada. No se recuperó ningún servicio.")

                return

            elif any(clave_recuperar == servicio[0] for servicio in servicios_cancelados):

                cursor.execute("SELECT * FROM Servicios WHERE idservicio = ?", (clave_recuperar,))

                servicio = cursor.fetchone()



                print("\nDetalle del servicio a recuperar:")

                print(f"Clave del servicio: {servicio[0]}")

                print(f"Nombre: {servicio[1]}")

                print(f"Costo: {servicio[2]:.2f}")



                confirmacion = input("¿Está seguro de que desea recuperar este servicio? (S/N): ").strip().upper()



                if confirmacion == 'S':

                    cursor.execute("UPDATE Servicios SET status = 1 WHERE idservicio = ?", (clave_recuperar,))

                    conn.commit()

                    print("El servicio ha sido recuperado exitosamente.")

                else:

                    print("La recuperación del servicio ha sido cancelada a solicitud del usuario.")

                break

            else:

                print("La clave ingresada no corresponde a un servicio cancelado en el sistema.")







def consultasreportes_servicios():

    while True:

        print("1. Búsqueda por clave de servicio")

        print("2. Búsqueda por nombre de servicio")

        print("3. Listado de servicios")

        print("4. Salir al menú Servicios")

        opcion = input("Seleccione una opción: ")

        if opcion == "1":   

            Busqueda_clave_servicio()

        elif opcion == "2":  

            Busqueda_nombre_servicio()

        elif opcion == "3":  

            Listado_servicios()

        elif opcion == "4":  

            print("Saliendo del programa. ¡Hasta luego!")

            break

        else:

            print("Opción no válida. Por favor, seleccione una opción válida.")



def Busqueda_clave_servicio():

    with sqlite3.connect("taller.db") as conn:

        cursor = conn.cursor()

        try:

            cursor.execute("SELECT idservicio, nombreservicio FROM Servicios WHERE status = 1")

            servicios = cursor.fetchall()



            if not servicios:

                print("No hay servicios registrados en la base de datos.")

                return



            print("Listado de servicios:")

            print("Clave  |  Nombre del Servicio")

            for servicio in servicios:

                print(f"{servicio[0]:<6} | {servicio[1]}")



            while True:

                try:

                    clave_servicio = int(input("Ingrese la clave del servicio que desea consultar: "))

                    if clave_servicio in [servicio[0] for servicio in servicios]:

                        break

                    else:

                        print("Clave no válida. Por favor, ingrese una clave válida.")

                except ValueError:

                    print("Error: La clave debe ser un número válido.")



            cursor.execute("SELECT idservicio, nombreservicio, costoservicio FROM Servicios WHERE idservicio = ?", (clave_servicio,))

            servicio_detalle = cursor.fetchone()



            print("\nDetalle del servicio:")

            print(f"Clave del Servicio: {servicio_detalle[0]}")

            print(f"Nombre del Servicio: {servicio_detalle[1]}")

            print(f"Costo del Servicio: {servicio_detalle[2]:.2f}")



        except sqlite3.Error as e:

            print(f"Error al buscar el servicio: {e}")



def Busqueda_nombre_servicio():

    with sqlite3.connect("taller.db") as conn:

        cursor = conn.cursor()

        try:

            nombre_servicio_buscar = input("Ingrese el nombre del servicio a buscar: ").strip().lower()



            cursor.execute("SELECT idservicio, nombreservicio, costoservicio FROM Servicios WHERE LOWER(nombreservicio) = ?", (nombre_servicio_buscar,))

            servicio_detalle = cursor.fetchone()



            if servicio_detalle:

                print("\nDetalle del servicio encontrado:")

                print(f"Clave del Servicio: {servicio_detalle[0]}")

                print(f"Nombre del Servicio: {servicio_detalle[1]}")

                print(f"Costo del Servicio: {servicio_detalle[2]:.2f}")

            else:

                print(f"No se encontró un servicio con el nombre '{nombre_servicio_buscar}'.")



        except sqlite3.Error as e:

            print(f"Error al buscar el servicio: {e}")

def Listado_servicios():
    while True:
        print("\nListado de servicios:")
        print("1. Ordenado por clave")
        print("2. Ordenado por nombre de servicio")
        print("3. Regresar al menú anterior")
        opcion = input("Seleccione una opción: ")

        if opcion == "1":  
            ordenar_servicios_por_clave()
        elif opcion == "2":  
            ordenar_servicios_por_nombre()
        elif opcion == "3": 
            print("Volviendo al menú de consultas y reportes de servicios.")
            break
        else:
            print("Opción no válida. Por favor, seleccione una opción válida.")

def ordenar_servicios_por_clave():
    with sqlite3.connect("taller.db") as conn:
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT idservicio, nombreservicio, costoservicio FROM Servicios WHERE status = 1 ORDER BY idservicio")
            servicios = cursor.fetchall()
            if servicios:
                print("\nListado de servicios ordenado por clave:")
                for servicio in servicios:
                    print(f"Clave: {servicio[0]}, Nombre: {servicio[1]}, Costo: {servicio[2]:.2f}")

                exportar_reporte_servicios(servicios, "ServiciosPorClave")

            else:
                print("No hay servicios registrados.")
        except sqlite3.Error as e:
            print(f"Error al obtener el listado de servicios: {e}")

def ordenar_servicios_por_nombre():
    with sqlite3.connect("taller.db") as conn:
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT idservicio, nombreservicio, costoservicio FROM Servicios WHERE status = 1 ORDER BY UPPER(nombreservicio)")
            servicios = cursor.fetchall()
            if servicios:
                print("\nListado de servicios ordenado por nombre:")
                for servicio in servicios:
                    print(f"Clave: {servicio[0]}, Nombre: {servicio[1]}, Costo: {servicio[2]:.2f}")

                exportar_reporte_servicios(servicios, "ServiciosPorNombre")
            else:
                print("No hay servicios registrados.")
        except sqlite3.Error as e:
            print(f"Error al obtener el listado de servicios: {e}")

def exportar_reporte_servicios(data, report_name):
    while True:
        opcion_exportar = input("\n¿Desea exportar el reporte a CSV, Excel o regresar al menú de reportes? (CSV/Excel/Regresar): ").strip().lower()
        if opcion_exportar == "csv":
            exportar_reporte_servicios_csv(data, report_name)
            break
        elif opcion_exportar == "excel":
            exportar_reporte_servicios_excel(data, report_name)
            break
        elif opcion_exportar == "regresar":
            break
        else:
            print("Opción no válida. Por favor, seleccione una opción válida.")

def exportar_reporte_servicios_csv(data, report_name):
    filename = f"{report_name}_{obtener_fecha_actual()}.csv"
    with open(filename, mode='w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(["Clave", "Nombre", "Costo"])
        for servicio in data:
            writer.writerow(servicio)
    print(f"El reporte ha sido exportado como {filename}")

def exportar_reporte_servicios_excel(data, report_name):
    filename = f"{report_name}_{obtener_fecha_actual()}.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = report_name

    bold = openpyxl.styles.Font(bold=True)
    bold_style = openpyxl.styles.NamedStyle(name="bold")
    bold_style.font = bold

    worksheet['A1'] = "Clave"
    worksheet['B1'] = "Nombre"
    worksheet['C1'] = "Costo"

    worksheet['A1'].style = bold_style
    worksheet['B1'].style = bold_style
    worksheet['C1'].style = bold_style

    row = 2

    for servicio in data:
        worksheet.cell(row=row, column=1, value=servicio[0])
        worksheet.cell(row=row, column=2, value=servicio[1])
        worksheet.cell(row=row, column=3, value=servicio[2])
        row += 1

    workbook.save(filename)
    print(f"El reporte ha sido exportado como {filename}")

def estadisticas():
    while True:
        mostrar_menu_estadisticas()
        opcion = input("Seleccione una opción: ")

        if opcion == "1":
            servicios_mas_prestados()
        elif opcion == "2":
            clientes_con_mas_notas()
        elif opcion == "3":
            promedio_montos_notas()
        elif opcion == "4":
            print("Volviendo al menú principal.")
            break
        else:
            print("Opción no válida. Por favor, seleccione una opción válida.")

def servicios_mas_prestados():
    print("\nReporte de Servicios Más Prestados")

    try:
        cantidad_servicios = int(input("Ingrese la cantidad de servicios más prestados a identificar: "))
        if cantidad_servicios <= 0:
            print("La cantidad de servicios debe ser un número positivo.")
            return
    except ValueError:
        print("Error: La cantidad de servicios debe ser un número entero.")
        return

    fecha_inicial = input("Ingrese la fecha inicial del período a reportar (mm/dd/yyyy): ")
    fecha_final = input("Ingrese la fecha final del período a reportar (mm/dd/yyyy): ")

    if not (validar_fechaestadistica(fecha_inicial) and validar_fechaestadistica(fecha_final)):
        print("Error: Las fechas ingresadas no tienen el formato correcto o la fecha final es anterior a la fecha inicial.")
        return

    with sqlite3.connect("taller.db") as conn:
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT idservicio FROM Notas WHERE fecha BETWEEN ? AND ?", (fecha_inicial, fecha_final))
            servicios_ids = cursor.fetchall()

            if not servicios_ids:
                print("No hay servicios registrados en el período especificado.")
                return

            contador_servicios = Counter(servicio[0] for servicio in servicios_ids)

            servicios_mas_prestados = contador_servicios.most_common(cantidad_servicios)

            print("\nReporte de Servicios Más Prestados:")
            print("{:<30} {:<10}".format("Nombre del Servicio", "Veces Prestado"))
            for servicio_id, cantidad in servicios_mas_prestados:
                cursor.execute("SELECT nombreservicio FROM Servicios WHERE idservicio = ?", (servicio_id,))
                nombre_servicio = cursor.fetchone()[0]
                print("{:<30} {:<10}".format(nombre_servicio, cantidad))

            while True:
                opcion_exportar = input("\n¿Desea exportar el reporte a CSV, Excel o regresar al menú de estadísticas? (CSV/Excel/Regresar): ").strip().lower()
                if opcion_exportar == "csv":
                    exportar_reporte_servicios_mas_prestados(conn, servicios_mas_prestados, fecha_inicial, fecha_final, "CSV")
                    break
                elif opcion_exportar == "excel":
                    exportar_reporte_servicios_mas_prestados(conn, servicios_mas_prestados, fecha_inicial, fecha_final, "Excel")
                    break
                elif opcion_exportar == "regresar":
                    break
                else:
                    print("Opción no válida. Por favor, seleccione una opción válida.")
        except sqlite3.Error as e:
            print(f"Error al generar el reporte de servicios más prestados: {e}")

def exportar_reporte_servicios_mas_prestados(conn, data, fecha_inicial, fecha_final, format):
    with sqlite3.connect("taller.db") as conn:
        cursor = conn.cursor()
        fecha_inicial_str = datetime.strptime(fecha_inicial, "%m/%d/%Y").strftime("%m_%d_%Y")
        fecha_final_str = datetime.strptime(fecha_final, "%m/%d/%Y").strftime("%m_%d_%Y")
        filename = f"ReporteServiciosMasPrestados_{fecha_inicial_str}_{fecha_final_str}.{format.lower()}"

        if format.lower() == "csv":
            with open(filename, mode='w', newline='') as file:
                writer = csv.writer(file)
                writer.writerow(["Nombre del Servicio", "Veces Prestado"])
                for servicio_id, cantidad in data:
                    cursor.execute("SELECT nombreservicio FROM Servicios WHERE idservicio = ?", (servicio_id,))
                    nombre_servicio = cursor.fetchone()[0]
                    writer.writerow([nombre_servicio, cantidad])
            print(f"El reporte ha sido exportado como {filename}")
        elif format.lower() == "excel":
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "ServiciosMasPrestados"

            bold = openpyxl.styles.Font(bold=True)
            bold_style = openpyxl.styles.NamedStyle(name="bold")
            bold_style.font = bold

            worksheet['A1'] = "Nombre del Servicio"
            worksheet['B1'] = "Veces Prestado"

            worksheet['A1'].style = bold_style
            worksheet['B1'].style = bold_style

            row = 2

            for servicio_id, cantidad in data:
                cursor.execute("SELECT nombreservicio FROM Servicios WHERE idservicio = ?", (servicio_id,))
                nombre_servicio = cursor.fetchone()[0]
                worksheet.cell(row=row, column=1, value=nombre_servicio)
                worksheet.cell(row=row, column=2, value=cantidad)
                row += 1

            workbook.save(filename)
            print(f"El reporte ha sido exportado como {filename}")

def clientes_con_mas_notas():
    print("\nReporte de Clientes con Más Notas")

    try:
        cantidad_clientes = int(input("Ingrese la cantidad de clientes con más notas a identificar: "))
        if cantidad_clientes <= 0:
            print("La cantidad de clientes debe ser un número positivo.")
            return
    except ValueError:
        print("Error: La cantidad de clientes debe ser un número entero.")
        return

    fecha_inicial = input("Ingrese la fecha inicial del período a reportar (mm/dd/yyyy): ")
    fecha_final = input("Ingrese la fecha final del período a reportar (mm/dd/yyyy): ")

    if not (validar_fechaestadistica(fecha_inicial) and validar_fechaestadistica(fecha_final)):
        print("Error: Las fechas ingresadas no tienen el formato correcto o la fecha final es anterior a la fecha inicial.")
        return

    with sqlite3.connect("taller.db") as conn:
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT idcliente FROM Notas WHERE fecha BETWEEN ? AND ?", (fecha_inicial, fecha_final))
            clientes_ids = cursor.fetchall()

            if not clientes_ids:
                print("No hay clientes registrados en el período especificado.")
                return

            contador_clientes = Counter(cliente[0] for cliente in clientes_ids)

            clientes_con_mas_notas = contador_clientes.most_common(cantidad_clientes)

            print("\nReporte de Clientes con Más Notas:")
            print("{:<30} {:<10}".format("Nombre del Cliente", "Notas Prestadas"))
            for cliente_id, cantidad in clientes_con_mas_notas:
                cursor.execute("SELECT nombrecliente FROM Clientes WHERE idcliente = ?", (cliente_id,))
                nombre_cliente = cursor.fetchone()[0]
                print("{:<30} {:<10}".format(nombre_cliente, cantidad))

            while True:
                opcion_exportar = input("\n¿Desea exportar el reporte a CSV, Excel o regresar al menú de estadísticas? (CSV/Excel/Regresar): ").strip().lower()
                if opcion_exportar == "csv":
                    exportar_reporte_clientes_con_mas_notas(conn, clientes_con_mas_notas, fecha_inicial, fecha_final, "CSV")
                    break
                elif opcion_exportar == "excel":
                    exportar_reporte_clientes_con_mas_notas(conn, clientes_con_mas_notas, fecha_inicial, fecha_final, "Excel")
                    break
                elif opcion_exportar == "regresar":
                    break
                else:
                    print("Opción no válida. Por favor, seleccione una opción válida.")
        except sqlite3.Error as e:
            print(f"Error al generar el reporte de clientes con más notas: {e}")

def exportar_reporte_clientes_con_mas_notas(conn, data, fecha_inicial, fecha_final, format):
    fecha_inicial_str = datetime.strptime(fecha_inicial, "%m/%d/%Y").strftime("%m_%d_%Y")
    fecha_final_str = datetime.strptime(fecha_final, "%m/%d/%Y").strftime("%m_%d_%Y")
    filename = f"ReporteClientesConMasNotas_{fecha_inicial_str}_{fecha_final_str}.{format.lower()}"

    cursor = conn.cursor()

    if format.lower() == "csv":
        with open(filename, mode='w', newline='') as file:
            writer = csv.writer(file)
            writer.writerow(["Nombre del Cliente", "Notas Prestadas"])
            for cliente_id, cantidad in data:
                cursor.execute("SELECT nombrecliente FROM Clientes WHERE idcliente = ?", (cliente_id,))
                nombre_cliente = cursor.fetchone()[0]
                writer.writerow([nombre_cliente, cantidad])
        print(f"El reporte ha sido exportado como {filename}")
    elif format.lower() == "excel":
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "ClientesConMasNotas"

        bold = openpyxl.styles.Font(bold=True)
        bold_style = openpyxl.styles.NamedStyle(name="bold")
        bold_style.font = bold

        worksheet['A1'] = "Nombre del Cliente"
        worksheet['B1'] = "Notas Prestadas"

        worksheet['A1'].style = bold_style
        worksheet['B1'].style = bold_style

        row = 2

        for cliente_id, cantidad in data:
            cursor.execute("SELECT nombrecliente FROM Clientes WHERE idcliente = ?", (cliente_id,))
            nombre_cliente = cursor.fetchone()[0]
            worksheet.cell(row=row, column=1, value=nombre_cliente)
            worksheet.cell(row=row, column=2, value=cantidad)
            row += 1

        workbook.save(filename)
        print(f"El reporte ha sido exportado como {filename}")

def promedio_montos_notas():
    try:
        with sqlite3.connect("taller.db") as conn:
            cursor = conn.cursor()

            fecha_inicial = input("Ingrese la fecha inicial del período a reportar (mm/dd/yyyy): ")
            fecha_final = input("Ingrese la fecha final del período a reportar (mm/dd/yyyy): ")

            if not (validar_fechaestadistica(fecha_inicial) and validar_fechaestadistica(fecha_final)):
                print("Error: Las fechas ingresadas no tienen el formato correcto o la fecha final es anterior a la fecha inicial.")
                return

            cursor.execute("""
                SELECT AVG(s.costoservicio) AS promedio_precios
                FROM Servicios s
                WHERE s.idservicio IN (
                    SELECT n.idservicio
                    FROM Notas n
                    WHERE n.fecha BETWEEN ? AND ?
                )
            """, (fecha_inicial, fecha_final))

            promedio_precios = cursor.fetchone()[0]

            if promedio_precios is not None:
                print(f"\nEl promedio de los precios de los servicios para el período especificado es: ${promedio_precios:.2f}")
            else:
                print("No hay servicios registrados en el período especificado.")

    except sqlite3.Error as e:
        print(f"Error al calcular el promedio de los precios de los servicios: {e}")

def obtener_fecha_actual():
    return datetime.now()

def validar_fechaestadistica(fecha):
    try:
        datetime.strptime(fecha, "%m/%d/%Y")
        return True
    except ValueError:
        return False


def validar_fecha(fecha):
    try:
        fecha_ingresada = datetime.strptime(fecha, '%d/%m/%Y')
        fecha_actual = obtener_fecha_actual()
        if fecha_ingresada <= fecha_actual:
            return True
        else:
            return False
    except ValueError:
        return False

def generar_clave_aleatoria():
    caracteres = string.ascii_uppercase + string.digits
    return ''.join(random.choice(caracteres) for _ in range(6))

def validar_rfc(rfc):
    rfc_fisica_pattern = r'^[A-Z]{4}\d{2}(0[1-9]|1[0-2])(0[1-9]|[1-2]\d|3[0-1])[A-Z\d]{3}$'
    rfc_moral_pattern = r'^[A-Z]{3}\d{2}(0[1-9]|1[0-2])(0[1-9]|[1-2]\d|3[0-1])[A-Z\d]{3}$'
    if re.match(rfc_fisica_pattern, rfc) or re.match(rfc_moral_pattern, rfc):
        return True
    else:
        return False

def validar_correo(correo):
    correo_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    if re.match(correo_pattern, correo):
        return True
    else:
        return False
    
def clavecliente_existe_(tabla, clave):
    with sqlite3.connect("taller.db") as conn:
        cursor = conn.cursor()
        try:
            cursor.execute(f"SELECT 1 FROM {tabla} WHERE idcliente = ?", (clave))
            return True
        except sqlite3.Error:
            return False

        
def claveservicio_existe_(tabla, clave):
    with sqlite3.connect("taller.db") as conn:
        cursor = conn.cursor()
        try:
            cursor.execute(f"SELECT 1 FROM {tabla} WHERE {tabla}idservicio = ?", (clave,))
            return cursor.fetchone() is not None
        except sqlite3.Error:
            return False
        
def mostrar_registros(tabla):
    print(f"\nRegistros de {tabla}")
    with sqlite3.connect("taller.db") as conn:
        cursor = conn.cursor()
        try:
            cursor.execute(f"SELECT * FROM {tabla}")
            rows = cursor.fetchall()
            for row in rows:
                print(row)
        except sqlite3.Error as e:
            print(f"Error al mostrar los registros de la tabla {tabla}: {e}")


def main():
    while True:
        mostrar_menu_principal()
        opcion = input("Seleccione una opción: ")
        if opcion == "1":    
            Notas()
        elif opcion == "2":  
            Clientes()
        elif opcion == "3":  
            Servicios()
        elif opcion == "4":  
            estadisticas()
        elif opcion == "5":  
            confirmacion = input("¿Está seguro de que desea salir del programa? (Sí/No): ").strip().lower()
            if confirmacion == "si" or confirmacion == "sí":
                print("Saliendo del programa. ¡Hasta luego!")
                break
            elif confirmacion == "no":
                print("Regresando al menú principal.")
            else:
                print("Respuesta no válida. Por favor, seleccione 'Sí' o 'No'.")
        else:
            print("Opción no válida. Por favor, seleccione una opción válida.")

def mostrar_menu_principal():
    print("\nMenú Principal:")
    print("1. Notas")
    print("2. Clientes")
    print("3. Servicios")
    print("4. Estadísticas")
    print("5. Salir")

if __name__ == "__main__":
    crear_tablas()
    main()

